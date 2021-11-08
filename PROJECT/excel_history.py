import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


def WriteIn(data, id):
    filepath = "HelloFlask/static/excel_history/RequestHistory_" + id + ".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    #массив с названиями столбцов
    mas_names = ["Номер запроса", "Номер способа", "Бизнес, привлекающий трафик", "Бизнес конкурентов",
                 "Радиус окружностей", "Адреса зафиксированных точек", "Координаты зафиксированных точек",
                 "Время сделанного запроса(UTC+3:00)"]

    #массив с индексами ячеек, где будут написаны названия столбцов
    mas_index = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    #цикл для заполнения/заливки названий столбцов
    for i in range(len(mas_names)):
        ws[mas_index[i]] = mas_names[i]
        ws[mas_index[i]].fill = PatternFill(start_color="FFDF59", fill_type="solid")

    #задаю конкретные длины столбцам
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 29
    ws.column_dimensions['D'].width = 19
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 31
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 35

    #задаю ширину границам ячеек с наименованиями столбцов
    for i in range(1, 9):
        ws.cell(row=1, column=i).border = thin_border

    #цикл для конеченого заполнения файла историей запросов
    for i in range(len(data.keys())):
        for j in range(1, 9):
            ws.cell(row=i + 2, column=j).border = thin_border

        ws["A" + str(i + 2)] = i + 1
        ws["B" + str(i + 2)] = int(data[str(i + 1)]["md_number"])
        ws["C" + str(i + 2)] = data[str(i + 1)]["your_business"]
        ws["D" + str(i + 2)] = data[str(i + 1)]["conc_business"]
        if data[str(i + 1)]["radius"] != "": ws["E" + str(i + 2)] = int(data[str(i + 1)]["radius"])
        ws["F" + str(i + 2)] = data[str(i + 1)]["stopped_points_adress"]
        ws["G" + str(i + 2)] = data[str(i + 1)]["stopped_points_coords"]
        ws["H" + str(i + 2)] = data[str(i + 1)]["time"]

    wb.save(filepath)
